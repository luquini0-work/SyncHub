"""
Academy USA — Sync script (cookie-based, no Selenium)
======================================================
Uses the Mindbody session cookie to download the ScheduleAtAGlance report
as Excel, then transforms it to CatchCorner CSV format.

How to get the cookie:
  1. Login at https://clients.mindbodyonline.com (business 610481)
  2. Go to Reports → Staff → Schedule at a Glance
  3. DevTools → Network → any request → copy Cookie header
  4. Paste into SyncHub dashboard → 🍪 Cookie on Academy USA

Run:
  python academy_usa_sync.py
  python academy_usa_sync.py --start 2026-05-01 --end 2026-07-01
  python academy_usa_sync.py --no-upload
"""

import os, re, csv, sys, io, datetime, argparse, requests

# ── Config ────────────────────────────────────────────────────────────────────

MB_BUSINESS_ID = "610481"
MB_REPORT_URL  = (
    "https://clients.mindbodyonline.com/app/business/Report/Staff"
    "/ScheduleAtAGlance/Generate"
)
MB_EXPORT_URL  = (
    "https://clients.mindbodyonline.com/app/business/Report/Staff"
    "/ScheduleAtAGlance/ExportToExcel"
)

CC_USERNAME    = os.environ.get("CC_MAIN_USER", "localloginformaint.en.a.nc.e.1.9.940@gmail.com")
CC_PASSWORD    = os.environ.get("CC_MAIN_PASS", "odM3ouk^YT4k6nmHmJN4ujXObSEH!5")
CC_FACILITY    = "2105"
CC_ACCESS_FROM = "Corporate"
SYNC_DAYS      = 60

# ── Fetch Excel from Mindbody ─────────────────────────────────────────────────

def get_cookie() -> str:
    cookie = os.environ.get("SYNC_COOKIE", "").strip()
    if not cookie:
        print("ERROR: No cookie found. Set SYNC_COOKIE env var or upload via SyncHub dashboard.")
        sys.exit(1)
    return cookie

def make_session(cookie: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Cookie": cookie,
        "Referer": "https://clients.mindbodyonline.com/",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s

def fetch_excel(start_date: datetime.date, end_date: datetime.date, cookie: str) -> bytes:
    """
    POST to the ScheduleAtAGlance ExportToExcel endpoint with the date range.
    Returns the raw Excel file bytes.
    """
    s = make_session(cookie)

    start_str = start_date.strftime("%m/%d/%Y")
    end_str   = end_date.strftime("%m/%d/%Y")

    print(f"[1/4] Fetching Excel from Mindbody ({start_str} → {end_str})...")

    # First hit the report page to get ASP.NET ViewState + cookies
    r = s.get(MB_REPORT_URL, params={"reportID": "undefined"}, timeout=30)
    if r.status_code in (401, 403) or "signin" in r.url.lower():
        print("ERROR: Cookie is expired or invalid. Please update the cookie via SyncHub dashboard.")
        sys.exit(1)

    # Extract ViewState and other hidden fields if present
    vs_match = re.search(r'name="__VIEWSTATE"\s+value="([^"]*)"', r.text)
    vsg_match = re.search(r'name="__VIEWSTATEGENERATOR"\s+value="([^"]*)"', r.text)
    ev_match = re.search(r'name="__EVENTVALIDATION"\s+value="([^"]*)"', r.text)
    rvt_match = re.search(r'name="__RequestVerificationToken"\s+value="([^"]*)"', r.text)

    form_data = {
        "requiredtxtDateStart": start_str,
        "requiredtxtDateEnd":   end_str,
        "exportToExcel":        "true",
    }
    if vs_match:
        form_data["__VIEWSTATE"] = vs_match.group(1)
    if vsg_match:
        form_data["__VIEWSTATEGENERATOR"] = vsg_match.group(1)
    if ev_match:
        form_data["__EVENTVALIDATION"] = ev_match.group(1)
    if rvt_match:
        form_data["__RequestVerificationToken"] = rvt_match.group(1)

    s.headers["Content-Type"] = "application/x-www-form-urlencoded"
    s.headers["Accept"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel, */*"

    # Try ExportToExcel endpoint directly
    resp = s.post(MB_EXPORT_URL, data=form_data, timeout=60)

    if resp.status_code == 200 and len(resp.content) > 1000 and (
        b"PK" in resp.content[:4] or b"\xd0\xcf" in resp.content[:4]
    ):
        print(f"      Got Excel: {len(resp.content)} bytes")
        return resp.content

    # Fallback: try the Generate endpoint with export param
    print(f"      ExportToExcel returned {resp.status_code}, trying Generate endpoint...")
    resp2 = s.post(MB_REPORT_URL, data={**form_data, "btnExcel": "Export to Excel"}, timeout=60)
    if resp2.status_code == 200 and len(resp2.content) > 1000:
        print(f"      Got Excel via Generate: {len(resp2.content)} bytes")
        return resp2.content

    print(f"ERROR: Could not download Excel (status {resp.status_code}).")
    print(f"Response preview: {resp.text[:300]}")
    print("The cookie may be expired. Please update it via SyncHub dashboard.")
    sys.exit(1)

# ── Transform ─────────────────────────────────────────────────────────────────

def _parse_date(val) -> str:
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        epoch = datetime.date(1899, 12, 30)
        return (epoch + datetime.timedelta(days=int(val))).strftime("%Y-%m-%d")
    try:
        s = str(val).strip()[:10]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    except Exception:
        pass
    return None

def _parse_time(val) -> str:
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    if isinstance(val, float):
        total_minutes = round(val * 24 * 60)
        h, m = divmod(total_minutes, 60)
        return f"{h % 24:02d}:{m:02d}"
    try:
        s = str(val).strip()
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%H:%M")
            except ValueError:
                pass
    except Exception:
        pass
    return None

def _make_resource(staff: str, notes: str) -> str:
    if not staff:
        return None
    staff = str(staff).strip()
    match = re.match(r"(\d+),\s*(.+)", staff)
    if not match:
        return None
    court_num = match.group(1)
    color     = match.group(2).strip()
    notes_str = str(notes).strip() if notes else ""
    is_hc = bool(re.search(r"HC", notes_str, re.IGNORECASE))
    is_fc = bool(re.search(r"FC", notes_str, re.IGNORECASE))
    if is_hc and not is_fc:
        return f"{court_num} {color}"
    return f"FC {color}"

def _fmt_date(s: str) -> str:
    try:
        return datetime.datetime.strptime(s[:10], "%Y-%m-%d").strftime("%m/%d/%Y")
    except Exception:
        return s

def _fmt_time(s: str) -> str:
    try:
        t = datetime.datetime.strptime(s[:5], "%H:%M")
        hour = t.strftime("%I").lstrip("0") or "12"
        return f"{hour}:{t.strftime('%M')} {t.strftime('%p')}"
    except Exception:
        return s

def transform_excel(excel_bytes: bytes) -> list:
    """Auto-detect XLS vs XLSX and parse accordingly."""
    # Detect format by magic bytes
    is_xls  = excel_bytes[:4] == b'\xd0\xcf\x11\xe0'  # OLE2 / Excel 97-2003
    is_xlsx = excel_bytes[:4] == b'PK\x03\x04'          # ZIP / Excel 2007+

    if is_xls:
        return _transform_xls(excel_bytes)
    elif is_xlsx:
        return _transform_xlsx(excel_bytes)
    else:
        # Try both
        try:
            return _transform_xls(excel_bytes)
        except Exception:
            return _transform_xlsx(excel_bytes)


def _get_col_indices(header_row):
    """Find column indices from header row values."""
    row_lower = [str(v).lower().strip() if v else "" for v in header_row]
    col_date  = row_lower.index("date") if "date" in row_lower else 0
    col_start = next((j for j, v in enumerate(row_lower) if "start" in v), 1)
    col_end   = next((j for j, v in enumerate(row_lower) if "end" in v), 2)
    col_staff = next((j for j, v in enumerate(row_lower) if "staff" in v), 4)
    col_notes = next((j for j, v in enumerate(row_lower) if "note" in v), 13)
    return col_date, col_start, col_end, col_staff, col_notes


def _transform_xls(excel_bytes: bytes) -> list:
    """Parse old .xls format using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=excel_bytes)

    # Find Input sheet
    sheet = None
    for name in wb.sheet_names():
        if name.lower() == "input":
            sheet = wb.sheet_by_name(name)
            break
    if sheet is None:
        sheet = wb.sheet_by_index(0)

    print(f"[2/4] Transforming XLS sheet '{sheet.name}' ({sheet.nrows} rows)...")

    # Find header row
    header_row_idx = 0
    col_date, col_start, col_end, col_staff, col_notes = 0, 1, 2, 4, 13
    for i in range(min(3, sheet.nrows)):
        vals = [str(sheet.cell_value(i, j)).lower().strip() for j in range(sheet.ncols)]
        if "date" in vals:
            header_row_idx = i
            col_date, col_start, col_end, col_staff, col_notes = _get_col_indices(
                [sheet.cell_value(i, j) for j in range(sheet.ncols)]
            )
            break

    rows = []
    for i in range(header_row_idx + 1, sheet.nrows):
        try:
            def cell(col):
                return sheet.cell(i, col) if sheet.ncols > col else None

            date_cell  = cell(col_date)
            start_cell = cell(col_start)
            end_cell   = cell(col_end)
            staff_val  = str(sheet.cell_value(i, col_staff)).strip() if sheet.ncols > col_staff else ""
            notes_val  = str(sheet.cell_value(i, col_notes)).strip() if sheet.ncols > col_notes else ""

            if not date_cell or date_cell.ctype == 0:
                continue

            # Parse date
            if date_cell.ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate_as_datetime(date_cell.value, wb.datemode)
                date_s = dt.strftime("%Y-%m-%d")
            elif date_cell.ctype == xlrd.XL_CELL_TEXT:
                s = date_cell.value.strip()[:10]
                date_s = None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        date_s = datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        pass
                if not date_s:
                    continue
            else:
                continue

            # Parse time
            def parse_xls_time(cell):
                if cell is None:
                    return None
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    return dt.strftime("%H:%M")
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    total_min = round(cell.value * 1440)
                    h, m = divmod(total_min, 60)
                    return f"{h%24:02d}:{m:02d}"
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    s = cell.value.strip()
                    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
                        try:
                            return datetime.datetime.strptime(s, fmt).strftime("%H:%M")
                        except ValueError:
                            pass
                return None

            start_s  = parse_xls_time(start_cell)
            end_s    = parse_xls_time(end_cell)
            resource = _make_resource(staff_val, notes_val)

            if date_s and start_s and end_s and resource:
                rows.append({"Date": date_s, "Start": start_s, "End": end_s, "Resource": resource})
        except Exception:
            pass

    rows.sort(key=lambda r: (r["Date"], r["Start"], r["Resource"]))
    print(f"      {len(rows)} rows transformed.")
    return rows


def _transform_xlsx(excel_bytes: bytes) -> list:
    """Parse new .xlsx format using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.lower() == "input"), wb.sheetnames[0])
    ws = wb[sheet_name]
    print(f"[2/4] Transforming XLSX sheet '{sheet_name}' ({ws.max_row} rows)...")

    header_row_idx = 1
    col_date, col_start, col_end, col_staff, col_notes = 0, 1, 2, 4, 13
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        row_lower = [str(v).lower().strip() if v else "" for v in row]
        if "date" in row_lower:
            header_row_idx = i
            col_date, col_start, col_end, col_staff, col_notes = _get_col_indices(row)
            break

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        date_val  = row[col_date]  if len(row) > col_date  else None
        start_val = row[col_start] if len(row) > col_start else None
        end_val   = row[col_end]   if len(row) > col_end   else None
        staff_val = row[col_staff] if len(row) > col_staff else None
        notes_val = row[col_notes] if len(row) > col_notes else None

        if date_val is None:
            continue

        date_s   = _parse_date(date_val)
        start_s  = _parse_time(start_val)
        end_s    = _parse_time(end_val)
        resource = _make_resource(staff_val, notes_val)

        if date_s and start_s and end_s and resource:
            rows.append({"Date": date_s, "Start": start_s, "End": end_s, "Resource": resource})

    rows.sort(key=lambda r: (r["Date"], r["Start"], r["Resource"]))
    print(f"      {len(rows)} rows transformed.")
    return rows

# ── Save CSV ──────────────────────────────────────────────────────────────────

def save_csv(rows: list, output_path: str = None) -> str:
    if not output_path:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(script_dir, f"academy_usa_output_{ts}.csv")

    formatted = [{"Date": _fmt_date(r["Date"]), "Start": _fmt_time(r["Start"]), "End": _fmt_time(r["End"]), "Resource": r["Resource"]} for r in rows]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Date", "Start", "End", "Resource"])
        writer.writeheader()
        writer.writerows(formatted)

    print(f"[3/4] CSV: {output_path} ({len(formatted)} rows)")
    if formatted:
        print(f"      Sample: {formatted[0]}")
    return output_path

# ── CatchCorner upload ────────────────────────────────────────────────────────

def upload_to_catchcorner(csv_path: str):
    print("[4/4] Uploading to CatchCorner...")
    h = {"Accept": "application/json, text/plain, */*", "Content-Type": "application/json",
         "Origin": "https://cc-stage-corporate.azurewebsites.net",
         "Referer": "https://cc-stage-corporate.azurewebsites.net/",
         "User-Agent": "Mozilla/5.0", "x-cc-platform": "1"}
    r = requests.Session().post(
        "https://www.catchcorner.com/api/shared/authentication/Login",
        json={"accessFrom": CC_ACCESS_FROM, "email": CC_USERNAME, "loginPlatform": 1, "password": CC_PASSWORD},
        headers=h, timeout=15)
    r.raise_for_status()
    token = r.json().get("access_token")
    if not token:
        raise ValueError("No CatchCorner token")
    print("      Login OK.")
    del h["Content-Type"]
    h["Authorization"] = f"Bearer {token}"
    with open(csv_path, "rb") as f:
        resp = requests.Session().post(
            f"https://www.catchcorner.com/api/back-office-shared/csvdump/dump/upload/0/{CC_FACILITY}/0",
            files={"file": (os.path.basename(csv_path), f, "multipart/form-data")},
            headers=h, timeout=60)
    if not resp.ok:
        print(f"      Upload failed {resp.status_code}: {resp.text[:200]}")
        resp.raise_for_status()
    print("      Uploaded OK.")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--start")
    parser.add_argument("--end")
    parser.add_argument("--output-file")
    parser.add_argument("--no-upload", action="store_true")
    parser.add_argument("--cookie-file", help="Path to file with cookie string")
    args = parser.parse_args()

    today      = datetime.date.today()
    start_date = datetime.date.fromisoformat(args.start) if args.start else today
    end_date   = datetime.date.fromisoformat(args.end)   if args.end   else today + datetime.timedelta(days=SYNC_DAYS)

    # Load cookie
    if args.cookie_file:
        with open(args.cookie_file, encoding="utf-8") as cf:
            os.environ["SYNC_COOKIE"] = cf.read().strip()

    print("=" * 55)
    print("  Academy USA Sync")
    print(f"  Range: {start_date} → {end_date} ({(end_date - start_date).days + 1} days)")
    print("=" * 55)

    cookie      = get_cookie()
    excel_bytes = fetch_excel(start_date, end_date, cookie)
    rows        = transform_excel(excel_bytes)

    if not rows:
        print("WARNING: No rows after transformation.")
        sys.exit(0)

    csv_path = save_csv(rows, args.output_file)

    if not args.no_upload:
        upload_to_catchcorner(csv_path)

    print(f"\nDone. {len(rows)} records synced.")

if __name__ == "__main__":
    main()